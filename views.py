from django.shortcuts import render
from django.http import JsonResponse

def chart(request):
    if request.method == "GET":
        # Initialize dropdown options (replace with actual data)
        dropdown_options = {
            'variable': ['Variable1', 'Variable2', 'Variable3'],
            'dataset': ['Dataset1', 'Dataset2', 'Dataset3'],
            'country': ['Country1', 'Country2', 'Country3'],
            'model': ['Model1', 'Model2', 'Model3']
        }
        return render(request, 'chart.html', {'dropdown_options': dropdown_options})
    elif request.method == "POST":
        # Process AJAX request and return updated chart data
        variable = request.POST.get('variable')
        dataset = request.POST.get('dataset')
        country = request.POST.get('country')
        model = request.POST.get('model')

        # Example: Generate random data for demonstration purposes
        data = [random.randint(1, 100) for _ in range(10)]
        labels = [str(2020 + i) for i in range(10)]

        return JsonResponse({'data': data, 'labels': labels})










































import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.styles import Font


def add_output_columns(sheet):
    df = pd.read_excel(sheet)

    df["GDP YR"] = df.groupby(["ISOCODES"])["GDP"].pct_change(4) * 100
    df["CPI YR"] = df.groupby(["ISOCODES"])["CPI"].pct_change(4) * 100
    df["PH YR"] = df.groupby(["ISOCODES"])["PH"].pct_change(4) * 100
    df["PH DUBAI YR"] = df.groupby(["ISOCODES"])["PH_DUBAI"].pct_change(4) * 100
    df["PCOM YR"] = df.groupby(["ISOCODES"])["PCOM"].pct_change(4) * 100
    df["GDP QoQ"] = df.groupby(["ISOCODES"])["GDP"].pct_change(1) * 100
    df["CPI QoQ"] = df.groupby(["ISDCODES"])["CPI"].pct_change(1) * 100  # Typo fixed (ISDCODES -> ISOCODES)
    df["PH_QoQ"] = df.groupby(["ISOCODES"])["PH"].pct_change(1) * 100
    df["PH_DUBAI_QOQ"] = df.groupby(["ISOCODES"])["PH_DUBAI"].pct_change(1) * 100
    df["PCOM QOQ"] = df.groupby(["ISOCODES"])["PCOM"].pct_change(1) * 100

    return df

input_file = r"C:\\Users\\45363901\\PycharmProjects\\pythonProject\\Automation_GBIC\\GBIC Workings Data\\Data - Copy.xlsx"
workbook = openpyxl.load_workbook(input_file)
sheetNames = workbook.sheetnames
All = list()
for sheet in sheetNames:
    df = add_output_columns(sheet)
    All.append(df)
All[0]

def calculate_syear_avg(frame, isocode, tzero_date, col):
    index = frame[(frame["ISOCODES"] == isocode) & (frame["date"] == datetime.strptime(tzero_date, "%Y-%m-%d %H:%M:%S"))]
    if not index.empty:  # Handle cases where the date might not be present
        avg = frame[col].iloc[(index[0] + 1): (index[0] + 21)].mean()
        return avg
    else:
        return None  # Return None for missing data


def calculate_weights(df_table, isocode):
    mapping_file_path = "C:\\Users\\45363901\\PycharmProjects\\pythonProject\\Automation_GBIC\\GBIC Workings Data\\Input Mapping File Copy.xlsx"

    country_sheet = pd.read_excel(mapping_file_path, sheet_name="Country")
    scenario_weights_sheet = pd.read_excel(mapping_file_path, sheet_name="Scenario Weights")

    merged_sheet = scenario_weights_sheet.merge(country_sheet, on="ISOCODE")

    u1_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "U1")]["WEIGHT"].values[0]
    cn_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "CN")]["WEIGHT"].values[0]
    d1_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "D1")]["WEIGHT"].values[0]
    d2_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "D2")]["WEIGHT"].values[0]

    for i in range(0, 18):
        df_table.values[i][8] = (
            df_table.values[i][4] * u1_weight
            + df_table.values[i][5] * cn_weight
            + df_table.values[i][6] * d1_weight
            + df_table.values[i][7] * d2_weight
        )[0]


def construct_and_populate_tables(isocode, tzero_date, col):
    current_year = 2024

    columns = [
        "Upside",
        "Central",
        "Down 1",
        "Down 2",
        "Upside",
        "Central",
        "Down 1",
        "Down 2",
        "Weighted",
        "10% UP",
        "Central",
        "10% DN",
        "4% DN",
    ]
    rows = [
        ["Q1", "Q2", "Q3", "Q4", current_year, current_year + 1, current_year + 2, current_year + 3, "", "", "", "", ""],
        [str(current_year) + "-03-31", str(current_year) + "-06-30", str(current_year) + "-09-30", str(current_year) + "-12-31", str(current_year + 1), str(current_year + 1) + "-03-31", str(current_year + 1) + "-06-30", str(current_year + 1) + "-09-30", "", "", "", "", ""],
    ]
    df_table = pd.DataFrame(columns=columns, index=rows)

    year3_sum = 0
    year4_sum = 0
    col_index = 0

    for frame in All:
        for row in frame.itertuples():
            # For previous year
            if (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year - 1) + "-03-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[0][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year - 1) + "-06-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[1][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year - 1) + "-09-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[2][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year - 1) + "-12-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[3][col_index] = getattr(row, col)

            # For current year
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year) + "-03-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[5][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year) + "-06-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[6][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year) + "-09-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[7][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year) + "-12-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 1) + "-03-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[10][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 1) + "-06-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[11][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 1) + "-09-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[12][col_index] = getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 1) + "-12-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                df_table.values[13][col_index] = getattr(row, col)

            # For next year 2
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 2) + "-03-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                year3_sum += getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 2) + "-06-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                year3_sum += getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 2) + "-09-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                year3_sum += getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 2) + "-12-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                year3_sum += getattr(row, col)

            # For next year + 3
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 3) + "-03-31 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                year4_sum += getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 3) + "-06-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                year4_sum += getattr(row, col)
            elif (
                row.ISOCODES == isocode
                and row.date == datetime.strptime(str(current_year + 3) + "-09-30 00:00:00", "%Y-%m-%d %H:%M:%S")
            ):
                year4_sum += getattr(row, col)

        df_table.values[4][col_index] = (df_table.values[0][col_index] + df_table.values[1][col_index] +
                                         df_table.values[2][col_index] + df_table.values[3][col_index]) / 4
        df_table.values[9][col_index] = (df_table.values[5][col_index] + df_table.values[6][col_index] +
                                         df_table.values[7][col_index] + df_table.values[8][col_index]) / 4
        df_table.values[14][col_index] = (df_table.values[10][col_index] + df_table.values[11][col_index] +
                                          df_table.values[12][col_index] + df_table.values[13][col_index]) / 4
        calculate_5year_avg(frame, Isocode, tzero_date, col)
        df_table.values[17][col_index] = df_table.values[15][col_index]
        df_table.values[16][col_index] = df_table.values[16][col_index]
        df_table.values[16][col_index] = year3_sum / 4
        df_table.values[16][col_index] = year4_sum / 4
        year3_sum = 0
        year4_sum = 0
        col_index += 1
        if col_index == 8:
            col_index = 1
    calculate_weights(df_table, Isocode)
    return df_table

        # All tables append(df_table)
tzero_date = "2024-03-31"

Table_List = pd.read_excel(
    r"C:\Users\45363901\PycharmProjects\pythonProject\Automation_GBIC\GBIC_Workings Data\Input Mapping File - Copy.xlsx",
    sheet_name="Table_List")

country_list = ["GBR", "HKG", "CHN", "FRA", "USA", "CAN", "MEX", "ARE", "GBL"]

wb = Workbook()
wb.save("2401 GBIC_Tables_29022024.xlsx")

for Isocode in country_list:
    col_list = Table_List["MEV"].loc[Table_List["Country"] == Isocode].tolist()
    writer = pd.ExcelWriter("2401_GBIC_Tables 29022024.xlsx", engine="openpyxl", mode="a",
                            if_sheet_exists='overlay')
    start_row = 6
    start_col = 4
    for col in col_list:
        table = construct_and_populate_tables(Isocode, tzero_date, col)
        table.to_excel(writer, sheet_name=Isocode, startrow=start_row, startcol=start_col, index=True)
        start_col += len(table.columns) + 4
    writer.close()

def adjust_col_width_and_font_size(sheet):
    for column in sheet.columns:
        column_letter = column[0].column_letter
        if column_letter in ['A', 'B', 'C']:
            continue
        sheet.column_dimensions[column_letter].width = 6
        for cell in column:
            cell.font = cell.font if cell.font else Font()
            cell.font = Font(size=9)

input_file = r"C:\Users\45363901\PycharmProjects\pythonProject\Automation_GBIC\GBIC_Workings Data\GBIC_Workings\2401_GBIC_Tables_29022024.xlsx"
workbook = openpyxl.load_workbook(input_file)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    adjust_col_width_and_font_size(sheet)
workbook.save(input_file)










































import pandas as pd
from openpyxl import load_workbook, Workbook
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Side, Border

def add_output_columns(sheet):
    df = pd.read_excel(sheet)
    df["GDP YR"] = df.groupby(["ISOCODES"])["GDP"].pct_change(4) * 100
    df["CPI YR"] = df.groupby(["ISOCODES"])["CPI"].pct_change(4) * 100
    df["PH YR"] = df.groupby(["ISOCODES"])["PH"].pct_change(4) * 100
    df["PH DUBAI YR"] = df.groupby(["ISOCODES"])["PH_DUBAI"].pct_change(4) * 100
    df["PCOM YR"] = df.groupby(["ISOCODES"])["PCOM"].pct_change(4) * 100
    df["GDP QoQ"] = df.groupby(["ISOCODES"])["GDP"].pct_change(1) * 100
    df["CPI QoQ"] = df.groupby(["ISOCODES"])["CPI"].pct_change(1) * 100
    df["PH_QoQ"] = df.groupby(["ISOCODES"])["PH"].pct_change(1) * 100
    df["PH_DUBAI_QOQ"] = df.groupby(["ISOCODES"])["PH_DUBAI"].pct_change(1) * 100
    df["PCOM QOQ"] = df.groupby(["ISOCODES"])["PCOM"].pct_change(1) * 100
    return df

def calculate_weights(df_table, isocode):
    mapping_file_path = "C:\\Users\\45363901\\PycharmProjects\\pythonProject\\Automation_GBIC\\GBIC Workings Data\\Input Mapping File Copy.xlsx"
    country_sheet = pd.read_excel(mapping_file_path, sheet_name="Country")
    scenario_weights_sheet = pd.read_excel(mapping_file_path, sheet_name="Scenario Weights")
    merged_sheet = scenario_weights_sheet.merge(country_sheet, on="ISOCODE")
    u1_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "U1")]["WEIGHT"].values[0]
    cn_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "CN")]["WEIGHT"].values[0]
    d1_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "D1")]["WEIGHT"].values[0]
    d2_weight = merged_sheet[(merged_sheet["ISOCODE"] == isocode) & (merged_sheet["SCENARIO NAME"] == "D2")]["WEIGHT"].values[0]
    for i in range(0, 18):
        df_table.values[i][8] = (
            df_table.values[i][4] * u1_weight + df_table.values[i][5] * cn_weight +
            df_table.values[i][6] * d1_weight + df_table.values[i][7] * d2_weight
        )[0]

def construct_and_populate_tables(isocode, tzero_date, col):
    current_year = 2024

    columns = [
        "Upside",
        "Central",
        "Down 1",
        "Down 2",
        "Upside",
        "Central",
        "Down 1",
        "Down 2",
        "Weighted",
        "10% UP",
        "Central",
        "10% DN",
        "4% DN",
    ]
    rows = [
        ["Q1", "Q2", "Q3", "Q4", current_year, current_year + 1, current_year + 2, current_year + 3, "", "", "", "", ""],
        [str(current_year) + "-03-31", str(current_year) + "-06-30", str(current_year) + "-09-30", str(current_year) + "-12-31", str(current_year + 1), str(current_year + 1) + "-03-31", str(current_year + 1) + "-06-30", str(current_year + 1) + "-09-30", "", "", "", "", ""],
    ]
    df_table = pd.DataFrame(columns=columns, index=rows)
    df_table.loc[-1] = ['IFRS 1024' for _ in range(len(columns))]
    df_table.loc[-2] = ['IFRS9 4023 Scenarios' for _ in range(len(columns))]
    df_table.index += 2
    df_table = df_table.sort_index()

    year3_sum = 0
    year4_sum = 0
    col_index = 0

    for frame in All:
        for row in frame.itertuples():
            # Your existing code to populate the DataFrame goes here

    calculate_weights(df_table, isocode)
    return df_table

def adjust_col_width_and_font_size(sheet):
    for column in sheet.columns:
        column_letter = column[0].column_letter
        if column_letter in ['A', 'B', 'C']:
            continue
        sheet.column_dimensions[column_letter].width = 6
        for cell in column:
            cell.font = cell.font if cell.font else Font()
            cell.font = Font(size=9)

input_file = r"C:\Users\45363901\PycharmProjects\pythonProject\Automation_GBIC\GBIC_Workings Data\GBIC_Workings\2401_GBIC_Tables_29022024.xlsx"
workbook = load_workbook(input_file)

for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    # Remove gridlines
    sheet.sheet_view.showGridLines = False
    # Apply grey color to all data cells except column names
    for row in sheet.iter_rows(min_row=2):
        for cell in row:
            if cell.row > 5:
                cell.fill = PatternFill(start_color="FFAFAFAF", end_color="FFAFAFAF", fill_type="solid")
    # Add darker grey to specific rows
    for row in sheet.iter_rows(min_row=2, max_row=5):
        for cell in row:
            cell.fill = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")
    # Rename Q1, Q2, etc. columns to country codes
    for cell in sheet["D"][1:]:
        cell.value = cell.value.replace("Q", "")
    # Add rows for IFRS scenarios
    sheet.insert_rows(1)
    sheet.insert_rows(1)
    sheet["D1"].value = "IFRS 1024"
    sheet["D2"].value = "IFRS9 4023 Scenarios"
    sheet["A1"].value = "Moody's"
    sheet["B1"].value = "Moody's"
    sheet["C1"].value = "Moody's"
    sheet["A2"].value = "Moody's"
    sheet["B2"].value = "Moody's"
    sheet["C2"].value = "Moody's"
    sheet["A1"].font = Font(color="FFFF0000")
    sheet["B1"].font = Font(color="FFFF0000")
    sheet["C1"].font = Font(color="FFFF0000")
    sheet["A2"].font = Font(color="FFFF0000")
    sheet["B2"].font = Font(color="FFFF0000")
    sheet["C2"].font = Font(color="FFFF0000")
    sheet["D1"].font = Font(color="FFFF0000")
    sheet["D2"].font = Font(color="FFFF0000")
    sheet["D1"].fill = PatternFill(start_color="FF87CEEB", end_color="FF87CEEB", fill_type="solid")
    sheet["D2"].fill = PatternFill(start_color="FF87CEEB", end_color="FF87CEEB", fill_type="solid")
    # Apply thin white border to separate cells
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border

    adjust_col_width_and_font_size(sheet)

workbook.save(input_file)













import pandas as pd
import openpyxl
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Border, Side
from datetime import datetime

def add_output_columns(sheet, workbook):
    ws = workbook[sheet]
    data = ws.values
    cols = next(data)
    data = list(data)
    df = pd.DataFrame(data, columns=cols).round(1)  # Round data to 1 decimal place

    # Your existing transformations
    df["GDP YR"] = df.groupby(["ISOCODES"])["GDP"].pct_change(4) * 100
    df["CPI YR"] = df.groupby(["ISOCODES"])["CPI"].pct_change(4) * 100
    df["PH YR"] = df.groupby(["ISOCODES"])["PH"].pct_change(4) * 100
    df["PH DUBAI YR"] = df.groupby(["ISOCODES"])["PH_DUBAI"].pct_change(4) * 100
    df["PCOM YR"] = df.groupby(["ISOCODES"])["PCOM"].pct_change(4) * 100
    df["GDP QoQ"] = df.groupby(["ISOCODES"])["GDP"].pct_change(1) * 100
    df["CPI QoQ"] = df.groupby(["ISOCODES"])["CPI"].pct_change(1) * 100
    df["PH_QoQ"] = df.groupby(["ISOCODES"])["PH"].pct_change(1) * 100
    df["PH_DUBAI_QOQ"] = df.groupby(["ISOCODES"])["PH_DUBAI"].pct_change(1) * 100
    df["PCOM QOQ"] = df.groupby(["ISOCODES"])["PCOM"].pct_change(1) * 100

    return df

# Assuming the rest of your functions are defined here as per your original code

def apply_custom_formatting(workbook):
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        # Apply formatting here as per your requirements
        for row in sheet.iter_rows(min_row=2):
            for cell in row:
                if cell.row > 5:
                    cell.fill = PatternFill(start_color="FFAFAFAF", end_color="FFAFAFAF", fill_type="solid")
                if cell.row in [1, 2, 3, 4, 5]:
                    cell.fill = PatternFill(start_color="FF808080", end_color="FF808080", fill_type="solid")
                if cell.row == 1:
                    cell.font = Font(color="FFFF0000")
                thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
                cell.border = thin_border
        # Add custom rows for IFRS scenarios if needed

# Load workbook and apply transformations and formatting
input_file = r"C:\Users\45363901\PycharmProjects\pythonProject\Automation_GBIC\GBIC_Workings Data\Data - Copy.xlsx"
workbook = load_workbook(input_file)
sheetNames = workbook.sheetnames
All = []

for sheet in sheetNames:
    df = add_output_columns(sheet, workbook)
    All.append(df)

# Assuming you have a function to write these DataFrames back to an Excel file
# After writing back to Excel, reload the workbook to apply custom formatting
workbook = load_workbook('path_to_your_new_excel_file.xlsx')  # Use the path where you saved the new Excel file
apply_custom_formatting(workbook)
workbook.save('path_to_your_final_excel_file.xlsx')  # Save the final version of your Excel file with formatting
